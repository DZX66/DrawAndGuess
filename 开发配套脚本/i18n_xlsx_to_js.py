import os
import openpyxl
import json
from collections import defaultdict

def xlsx_to_js(input_path, output_path):
    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True)
    except Exception as e:
        print(f"文件加载失败: {str(e)}")
        return

    # 获取第一个工作表
    sheet = wb.active
    print(wb.sheetnames)
    if not sheet:
        print("Excel文件中没有有效的工作表")
        return

    # 获取有效标题行
    header = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        header = [cell.value for cell in row if cell.value]
        if header:  # 找到第一个非空行作为标题
            break

    if not header or header[0].lower() != 'key':
        print("第一列必须是'key'标题")
        return

    languages = header[1:]  # 排除key列后的语言列表

    # 构建翻译字典
    translations = defaultdict(lambda: defaultdict(dict))

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        key_cell = row[0]
        if not key_cell or not key_cell.value:
            continue

        key = str(key_cell.value).strip()
        key_parts = [key]
        
        for lang_idx, lang in enumerate(languages, start=1):
            if lang_idx >= len(row):
                continue
                
            cell = row[lang_idx]
            if not cell or not cell.value:
                continue

            try:
                cell_value = str(cell.value).strip()
            except:
                print(f"格式错误 @ 行{row_idx} {lang}.{key}")
                continue

            # 递归创建嵌套结构
            current = translations[lang]
            for part in key_parts[:-1]:
                current = current.setdefault(part, {})
            
            # 设置最终值
            current[key_parts[-1]] = cell_value

    # 生成JS文件
    js_content = "const translations = {\n"
    for lang_idx, (lang, trans_dict) in enumerate(translations.items()):
        lang_data = json.dumps(trans_dict, ensure_ascii=False, indent=2)
        js_content += f'  "{lang}": {lang_data}'
        js_content += ",\n" if lang_idx < len(translations)-1 else "\n"
    js_content += "};\n"

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

if __name__ == '__main__':
    os.chdir(os.path.dirname(__file__))
    xlsx_to_js(
        input_path='translations.xlsx',
        output_path='i18n.js'
    )